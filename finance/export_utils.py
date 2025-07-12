"""
Utilities for exporting reports to PDF and Excel formats
"""

import io
from datetime import datetime
from decimal import Decimal
from django.http import HttpResponse
from django.template.loader import get_template
from django.template import Context
from django.conf import settings

# Excel libraries
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF libraries
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class ReportExporter:
    """
    Class to handle export of reports to different formats
    """
    
    def __init__(self, report_data, report_type, title="Báo cáo"):
        self.report_data = report_data
        self.report_type = report_type
        self.title = title
        self.created_date = datetime.now()
    
    def export_to_excel(self):
        """
        Export report to Excel format
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo cáo"
        
        # Style definitions
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = self.title
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        
        # Date
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Ngày tạo: {self.created_date.strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].alignment = center_alignment
        
        # Add content based on report type
        if self.report_type == 'sales':
            self._add_sales_content_excel(ws)
        elif self.report_type == 'debt':
            self._add_debt_content_excel(ws)
        elif self.report_type == 'inventory':
            self._add_inventory_content_excel(ws)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.title}_{self.created_date.strftime("%Y%m%d")}.xlsx"'
        return response
    
    def _add_sales_content_excel(self, ws):
        """Add sales report content to Excel"""
        # Headers
        headers = ['Mã đại lý', 'Tên đại lý', 'Số phiếu xuất', 'Tổng doanh số', 'Tỷ lệ (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        sales_data = self.report_data.get('sales', [])
        total_sales = sum(item.get('total_sales', 0) for item in sales_data)
        
        for row, item in enumerate(sales_data, 5):
            ws.cell(row=row, column=1, value=f"DL{item.get('agency_id', '')}")
            ws.cell(row=row, column=2, value=item.get('agency_name', ''))
            ws.cell(row=row, column=3, value=item.get('total_issues', 0))
            ws.cell(row=row, column=4, value=item.get('total_sales', 0))
            
            # Calculate percentage
            percentage = (item.get('total_sales', 0) / total_sales * 100) if total_sales > 0 else 0
            ws.cell(row=row, column=5, value=f"{percentage:.1f}%")
    
    def _add_debt_content_excel(self, ws):
        """Add debt report content to Excel"""
        # Headers
        headers = ['Mã đại lý', 'Tên đại lý', 'Nợ đầu kỳ', 'Phát sinh', 'Thanh toán', 'Nợ cuối kỳ']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        debt_data = self.report_data.get('summary', [])
        
        for row, item in enumerate(debt_data, 5):
            ws.cell(row=row, column=1, value=f"DL{item.get('agency_id', '')}")
            ws.cell(row=row, column=2, value=item.get('agency_name', ''))
            ws.cell(row=row, column=3, value=item.get('debt_beginning', 0))
            ws.cell(row=row, column=4, value=item.get('debt_incurred', 0))
            ws.cell(row=row, column=5, value=item.get('debt_paid', 0))
            ws.cell(row=row, column=6, value=item.get('debt_ending', 0))
    
    def _add_inventory_content_excel(self, ws):
        """Add inventory report content to Excel"""
        # Headers
        headers = ['Mã hàng', 'Tên hàng', 'Đơn vị', 'Tồn kho', 'Đơn giá', 'Thành tiền']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        items_data = self.report_data.get('items', [])
        
        for row, item in enumerate(items_data, 5):
            ws.cell(row=row, column=1, value=item.get('item_id', ''))
            ws.cell(row=row, column=2, value=item.get('item_name', ''))
            ws.cell(row=row, column=3, value=item.get('unit_name', ''))
            ws.cell(row=row, column=4, value=item.get('stock_quantity', 0))
            ws.cell(row=row, column=5, value=item.get('price', 0))
            ws.cell(row=row, column=6, value=item.get('total_value', 0))
        
        # Total row
        total_row = len(items_data) + 6
        ws.cell(row=total_row, column=5, value="Tổng cộng:").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=self.report_data.get('total_value', 0)).font = Font(bold=True)
    
    def export_to_pdf(self):
        """
        Export report to PDF format
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Build story
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(self.title, title_style))
        story.append(Paragraph(f"Ngày tạo: {self.created_date.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Add content based on report type
        if self.report_type == 'sales':
            self._add_sales_content_pdf(story, styles)
        elif self.report_type == 'debt':
            self._add_debt_content_pdf(story, styles)
        elif self.report_type == 'inventory':
            self._add_inventory_content_pdf(story, styles)
        
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.title}_{self.created_date.strftime("%Y%m%d")}.pdf"'
        return response
    
    def _add_sales_content_pdf(self, story, styles):
        """Add sales report content to PDF"""
        # Data
        sales_data = self.report_data.get('sales', [])
        total_sales = sum(item.get('total_sales', 0) for item in sales_data)
        
        # Table data
        table_data = [['Mã đại lý', 'Tên đại lý', 'Số phiếu xuất', 'Tổng doanh số', 'Tỷ lệ (%)']]
        
        for item in sales_data:
            percentage = (item.get('total_sales', 0) / total_sales * 100) if total_sales > 0 else 0
            table_data.append([
                f"DL{item.get('agency_id', '')}",
                item.get('agency_name', ''),
                str(item.get('total_issues', 0)),
                f"{item.get('total_sales', 0):,.0f}",
                f"{percentage:.1f}%"
            ])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    def _add_debt_content_pdf(self, story, styles):
        """Add debt report content to PDF"""
        # Data
        debt_data = self.report_data.get('summary', [])
        
        # Table data
        table_data = [['Mã đại lý', 'Tên đại lý', 'Nợ đầu kỳ', 'Phát sinh', 'Thanh toán', 'Nợ cuối kỳ']]
        
        for item in debt_data:
            table_data.append([
                f"DL{item.get('agency_id', '')}",
                item.get('agency_name', ''),
                f"{item.get('debt_beginning', 0):,.0f}",
                f"{item.get('debt_incurred', 0):,.0f}",
                f"{item.get('debt_paid', 0):,.0f}",
                f"{item.get('debt_ending', 0):,.0f}"
            ])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    def _add_inventory_content_pdf(self, story, styles):
        """Add inventory report content to PDF"""
        # Data
        items_data = self.report_data.get('items', [])
        
        # Table data
        table_data = [['Mã hàng', 'Tên hàng', 'Đơn vị', 'Tồn kho', 'Đơn giá', 'Thành tiền']]
        
        for item in items_data:
            table_data.append([
                str(item.get('item_id', '')),
                item.get('item_name', ''),
                item.get('unit_name', ''),
                str(item.get('stock_quantity', 0)),
                f"{item.get('price', 0):,.0f}",
                f"{item.get('total_value', 0):,.0f}"
            ])
        
        # Add total row
        table_data.append(['', '', '', '', 'Tổng cộng:', f"{self.report_data.get('total_value', 0):,.0f}"])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table) 